import os
import shutil
from datetime import datetime, date
from io import BytesIO

from flask import (
    Flask, render_template, request, redirect, url_for, flash, send_file, session
)
from flask_login import (
    LoginManager, login_user, login_required, logout_user, current_user
)
from werkzeug.security import generate_password_hash, check_password_hash

from flask_migrate import Migrate  # DODAJ TEN IMPORT!

from forms import (
    LoginForm, RegisterForm, EmployeeForm, FieldForm, BerryVarietyForm,
    WorkTypeForm, DailyHarvestForm, EntryForm, PresenceForm
)
import openpyxl

# Tylko import modeli!
from models import db, User, Employee, Field, BerryVariety, WorkType, DailyHarvest, Entry, Presence

app = Flask(__name__)
app.config['SECRET_KEY'] = 'bardzo-tajny'
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///borowki.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

login_manager = LoginManager(app)
login_manager.login_view = 'login'

migrate = Migrate(app, db)

@app.context_processor
def inject_now():
    return {'current_year': datetime.now().year}

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

with app.app_context():
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin')
        admin.set_password('admin')
        db.session.add(admin)
        db.session.commit()

@app.context_processor
def inject_now():
    return {'current_year': datetime.now().year}

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

with app.app_context():
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin')
        admin.set_password('admin')
        db.session.add(admin)
        db.session.commit()

@app.route('/rozliczenie_czasu_pracy', methods=['GET'])
@login_required
def rozliczenie_czasu_pracy():
    start_date = request.args.get('start_date', date.today().isoformat())
    end_date = request.args.get('end_date', date.today().isoformat())

    employees = Employee.query.filter_by(is_active=True).order_by(Employee.name).all()
    presences = Presence.query.filter(Presence.date >= start_date, Presence.date <= end_date).all()
    entries = Entry.query.filter(Entry.date >= start_date, Entry.date <= end_date).all()
    harvests = DailyHarvest.query.filter(DailyHarvest.date >= start_date, DailyHarvest.date <= end_date).all()
    # TYLKO pola zbierane w zakresie dat:
    fields = (
        Field.query
        .join(DailyHarvest, Field.id == DailyHarvest.field_id)
        .filter(DailyHarvest.date >= start_date, DailyHarvest.date <= end_date)
        .distinct()
        .all()
    )

    data = []
    for emp in employees:
        presence = next((p for p in presences if p.employee_id == emp.id), None)
        time_in = presence.time_in if presence else None
        time_out = presence.time_out if presence else None
        break_minutes = presence.break_minutes if presence and presence.break_minutes else 0

        hours = sum(e.hours or 0 for e in entries if e.employee_id == emp.id and (e.hours or 0) > 0)
        total_kg = sum(h.quantity_kg for h in harvests if h.employee_id == emp.id)
        field_weights = {f.name: sum(h.quantity_kg for h in harvests if h.employee_id == emp.id and h.field_id == f.id) for f in fields}
        total_sztuki = sum(e.quantity or 0 for e in entries if e.employee_id == emp.id and e.work_type and e.work_type.unit == "sztuki")

        piece_pay = sum((h.quantity_kg or 0) * (emp.piece_rate or 0) for h in harvests if h.employee_id == emp.id)
        hourly_pay = hours * (emp.hourly_rate or 0)
        suma_czasu = ""
        if time_in and time_out:
            delta = datetime.combine(date.today(), time_out) - datetime.combine(date.today(), time_in)
            total_minutes = delta.total_seconds() // 60 - (break_minutes or 0)
            godziny = int(total_minutes // 60)
            minuty = int(total_minutes % 60)
            suma_czasu = f"{godziny:02d}:{minuty:02d}"

        data.append({
            'kod_pr': emp.external_id or emp.id,
            'imie': emp.name.split()[0] if emp.name else '',
            'nazwisko': ' '.join(emp.name.split()[1:]) if emp.name else '',
            'start': time_in.strftime('%H:%M') if time_in else '',
            'przerwa': f"{break_minutes:02d}:00" if break_minutes else '00:00',
            'stop': time_out.strftime('%H:%M') if time_out else '',
            'praca_akord': f"{total_kg:.2f}",
            'praca_godziny': f"{hours:.2f}",
            'suma_czasu': suma_czasu,
            'field_weights': field_weights,
            'sztuki_razem': total_sztuki,
            'razem_waga': f"{total_kg:.2f}",
            'placa_akord': f"{piece_pay:.2f}",
            'placa_godziny': f"{hourly_pay:.2f}",
            'wynagrodzenie_brutto': f"{(piece_pay + hourly_pay):.2f}"
        })

    return render_template('rozliczenie_czasu_pracy.html', data=data, fields=fields, start_date=start_date, end_date=end_date)

@app.route('/presence', methods=['GET', 'POST'])
@login_required
def presence():
    form = PresenceForm()
    employees = Employee.query.order_by(Employee.name).all()
    form.employee_id.choices = [(e.id, e.name) for e in Employee.query.order_by(Employee.name)]
    if form.validate_on_submit():
        presence = Presence(
            employee_id=form.employee_id.data,
            date=form.date.data,
            time_in=form.time_in.data,
            time_out=form.time_out.data,
            comment=form.comment.data
        )
        db.session.add(presence)
        db.session.commit()
        flash("Zapisano obecność.", "success")
        return redirect(url_for('presence_list'))
    return render_template('presence.html', form=form)

@app.route('/presence_list')
@login_required
def presence_list():
    presences = Presence.query.order_by(Presence.date.desc(), Presence.time_in.desc()).all()
    return render_template('presence_list.html', presences=presences)

# --- POPRAWIONY I JEDYNY ENDPOINT FAST HARVEST ---
@app.route('/fast_harvest', methods=['GET', 'POST'])
@login_required
def fast_harvest():
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.name).all()
    today = date.today()
    kg_data = session.get('fast_harvest_kg_data', {})

    if request.method == 'POST':
        action = request.form.get('action')
        # Szybkie dodanie koszyka wybranemu pracownikowi przez Select2
        if action == 'add_selected':
            emp_id = request.form.get('selected_employee_id')
            kg = request.form.get('selected_employee_kg')
            if emp_id and kg:
                try:
                    kg_val = float(kg)
                    if kg_val > 0:
                        emp_id_str = str(emp_id)
                        if emp_id_str not in kg_data:
                            kg_data[emp_id_str] = []
                        kg_data[emp_id_str].append(kg_val)
                        session['fast_harvest_kg_data'] = kg_data
                        flash('Dodano koszyk!', 'success')
                    else:
                        flash('Podaj wagę większą od 0!', 'danger')
                except ValueError:
                    flash('Błędna wartość kg.', 'danger')
            else:
                flash('Wybierz pracownika i podaj ilość!', 'danger')
        # Zbiorczy wpis z tabeli
        elif action == 'add_bulk':
            for emp in employees:
                kg_str = request.form.get(f'kg_{emp.id}')
                if kg_str:
                    try:
                        kg_val = float(kg_str)
                        if kg_val > 0:
                            emp_id_str = str(emp.id)
                            if emp_id_str not in kg_data:
                                kg_data[emp_id_str] = []
                            kg_data[emp_id_str].append(kg_val)
                            session['fast_harvest_kg_data'] = kg_data
                    except ValueError:
                        flash(f'Nieprawidłowa liczba dla {emp.name}', 'danger')
            flash('Dodano koszyki!', 'success')
        # Usuwanie koszyka
        elif 'remove_emp_id' in request.form and 'remove_idx' in request.form:
            emp_id = request.form.get('remove_emp_id')
            idx = int(request.form.get('remove_idx'))
            if emp_id in kg_data and 0 <= idx < len(kg_data[emp_id]):
                kg_data[emp_id].pop(idx)
                session['fast_harvest_kg_data'] = kg_data
                flash('Usunięto koszyk.', 'info')
        return redirect(url_for('fast_harvest'))

    # Przygotuj podsumowanie do tabeli
    summary = []
    for emp in employees:
        emp_id = str(emp.id)
        details = list(enumerate(kg_data.get(emp_id, [])))
        total = round(sum(kg_data.get(emp_id, [])), 2)
        summary.append({'id': emp_id, 'name': emp.name, 'total': total, 'details': details})

    return render_template('fast_harvest.html', employees=employees, summary=summary)

@app.route('/fast_harvest_finish', methods=['GET', 'POST'])
@login_required
def fast_harvest_finish():
    today = date.today()
    kg_data = session.get('fast_harvest_kg_data', {})
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.name).all()
    varieties = BerryVariety.query.all()
    fields = Field.query.all()

    if not kg_data:
        flash("Najpierw dodaj koszyki!", "warning")
        return redirect(url_for('fast_harvest'))

    if request.method == 'POST':
        piece_rate = float(request.form['piece_rate'])
        variety_id = int(request.form['variety_id'])
        field_id = int(request.form['field_id'])

        for emp_id, kg_list in kg_data.items():
            total_kg = round(sum(kg_list), 2)
            if total_kg > 0:
                harvest = DailyHarvest(
                    date=today,
                    employee_id=int(emp_id),
                    quantity_kg=total_kg,
                    variety_id=variety_id,
                    field_id=field_id,
                    comment=f"Szybki wpis, sum koszyków: {kg_list}"
                )
                db.session.add(harvest)
                worktype = WorkType.query.filter_by(is_piece_rate=True).first()
                if worktype:
                    entry = Entry(
                        date=today,
                        employee_id=int(emp_id),
                        work_type_id=worktype.id,
                        hours=0,
                        quantity=total_kg,
                        variety_id=variety_id,
                        field_id=field_id,
                        comment=f"Automatyczny wpis z sumy koszyków: {kg_list}",
                        piece_rate=piece_rate
                    )
                    db.session.add(entry)
        db.session.commit()
        session.pop('fast_harvest_kg_data', None)
        flash("Zapisano sumy zbiorów!", "success")
        return redirect(url_for('harvests'))

    summary = []
    for emp in employees:
        emp_id = str(emp.id)
        details = list(enumerate(kg_data.get(emp_id, [])))
        total = round(sum(kg_data.get(emp_id, [])), 2)
        summary.append({'id': emp_id, 'name': emp.name, 'total': total, 'details': details})

    return render_template('fast_harvest_finish.html', varieties=varieties, fields=fields, summary=summary)

# ---- RESZTA TWOICH ENDPOINTÓW ----

@app.route('/download_backup', methods=['POST'])
@login_required
def download_backup():
    if current_user.username != 'Admin':
        flash('Brak dostępu!', 'danger')
        return redirect(url_for('dashboard'))
    db_path = app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
    backup_path = db_path + ".backup"
    shutil.copy2(db_path, backup_path)
    response = send_file(
        backup_path,
        as_attachment=True,
        download_name='borowki_backup.sqlite',
        mimetype='application/x-sqlite3'
    )
    @response.call_on_close
    def cleanup():
        try:
            os.remove(backup_path)
        except Exception:
            pass
    return response

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data.strip()).first()
        if user and user.check_password(form.password.data):
            if not user.is_approved:
                flash('Twoje konto oczekuje na zatwierdzenie przez administratora.', 'warning')
                return redirect(url_for('login'))
            login_user(user)
            flash('Zalogowano!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Błędna nazwa użytkownika lub hasło.', 'danger')
    return render_template('login.html', form=form)

@app.route('/pending_users')
@login_required
def pending_users():
    if current_user.username != 'admin':
        flash('Brak dostępu!', 'danger')
        return redirect(url_for('dashboard'))
    users = User.query.filter_by(is_approved=False).all()
    return render_template('pending_users.html', users=users)

@app.route('/approve_user/<int:user_id>', methods=['POST'])
@login_required
def approve_user(user_id):
    if current_user.username != 'admin':
        flash('Brak dostępu!', 'danger')
        return redirect(url_for('dashboard'))
    user = User.query.get_or_404(user_id)
    user.is_approved = True
    db.session.commit()
    flash(f'Zatwierdzono użytkownika {user.username}', 'success')
    return redirect(url_for('pending_users'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    form = RegisterForm()
    if form.validate_on_submit():
        if User.query.filter_by(username=form.username.data.strip()).first():
            flash('Taki użytkownik już istnieje.', 'danger')
        else:
            user = User(username=form.username.data.strip())
            user.set_password(form.password.data)
            db.session.add(user)
            db.session.commit()
            flash('Konto utworzone. Możesz się zalogować.', 'success')
            return redirect(url_for('login'))
    return render_template('register.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Wylogowano.', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

# --- PRACOWNICY ---
@app.route('/employees')
@login_required
def employees():
    # Obsługa filtrów (jeśli chcesz, możesz rozbudować)
    q = request.args.get('q', '').strip()
    team = request.args.get('team', '')
    status = request.args.get('status', '')
    selected_id = request.args.get('selected_id', type=int)

    # Zespoły do filtrów
    teams = db.session.query(Employee.team).distinct().all()
    teams = [t[0] for t in teams if t[0]]

    # Podstawowe zapytanie
    query = Employee.query

    # Filtrowanie po nazwie/pracowniku
    if q:
        query = query.filter(Employee.name.ilike(f"%{q}%"))

    # Filtrowanie po zespole
    if team:
        query = query.filter(Employee.team == team)

    # Filtrowanie po statusie
    if status == 'active':
        query = query.filter(Employee.is_active == True)
    elif status == 'inactive':
        query = query.filter(Employee.is_active == False)

    employees = query.order_by(Employee.name).all()

    # Pobierz wybranego pracownika (do szczegółów po prawej)
    employee = None
    if selected_id:
        employee = Employee.query.get(selected_id)

    return render_template(
        'employees.html',
        employees=employees,
        employee=employee,
        selected_id=selected_id,
        teams=teams
    )

@app.route('/deactivate_employee/<int:employee_id>', methods=['POST'])
@login_required
def deactivate_employee(employee_id):
    emp = Employee.query.get_or_404(employee_id)
    emp.is_active = False
    db.session.commit()
    flash('Pracownik został dezaktywowany.', 'info')
    return redirect(url_for('employees'))

@app.route('/activate_employee/<int:employee_id>', methods=['POST'])
@login_required
def activate_employee(employee_id):
    emp = Employee.query.get_or_404(employee_id)
    emp.is_active = True
    db.session.commit()
    flash('Pracownik został ponownie aktywowany.', 'success')
    return redirect(url_for('employees'))

@app.route('/add_employee', methods=['GET', 'POST'])
@login_required
def add_employee():
    form = EmployeeForm()
    if form.validate_on_submit():
        emp = Employee(
            external_id=form.external_id.data or None,  # Dodaj obsługę external_id!
            name=form.name.data,
            hourly_rate=float(form.hourly_rate.data),
            piece_rate=float(form.piece_rate.data),
            is_active=form.is_active.data
        )
        db.session.add(emp)
        db.session.commit()
        flash('Dodano pracownika!', 'success')
        return redirect(url_for('employees'))
    return render_template('add_employee.html', form=form)

@app.route('/edit_employee/<int:employee_id>', methods=['GET', 'POST'])
@login_required
def edit_employee(employee_id):
    emp = Employee.query.get_or_404(employee_id)
    form = EmployeeForm(obj=emp)
    if form.validate_on_submit():
        emp.external_id = form.external_id.data or None
        emp.name = form.name.data
        emp.hourly_rate = float(form.hourly_rate.data)
        emp.piece_rate = float(form.piece_rate.data)
        emp.is_active = form.is_active.data
        db.session.commit()
        flash('Zaktualizowano pracownika!', 'success')
        return redirect(url_for('employees'))
    return render_template('edit_employee.html', form=form, employee=emp)

@app.route('/delete_employee/<int:employee_id>', methods=['POST'])
@login_required
def delete_employee(employee_id):
    emp = Employee.query.get_or_404(employee_id)
    db.session.delete(emp)
    db.session.commit()
    flash('Usunięto pracownika.', 'success')
    return redirect(url_for('employees'))

# --- POLA ---
@app.route('/fields')
@login_required
def fields():
    fields = Field.query.all()
    return render_template('fields.html', fields=fields)

@app.route('/add_field', methods=['GET', 'POST'])
@login_required
def add_field():
    form = FieldForm()
    if form.validate_on_submit():
        f = Field(name=form.name.data)
        db.session.add(f)
        db.session.commit()
        flash('Dodano pole.', 'success')
        return redirect(url_for('fields'))
    return render_template('add_field.html', form=form)

@app.route('/edit_field/<int:field_id>', methods=['GET', 'POST'])
@login_required
def edit_field(field_id):
    field = Field.query.get_or_404(field_id)
    form = FieldForm(obj=field)
    if form.validate_on_submit():
        field.name = form.name.data
        db.session.commit()
        flash('Zaktualizowano pole.', 'success')
        return redirect(url_for('fields'))
    return render_template('edit_field.html', form=form, field=field)

@app.route('/delete_field/<int:field_id>', methods=['POST'])
@login_required
def delete_field(field_id):
    field = Field.query.get_or_404(field_id)
    db.session.delete(field)
    db.session.commit()
    flash('Usunięto pole.', 'success')
    return redirect(url_for('fields'))

# --- ODMIANY ---
@app.route('/varieties')
@login_required
def varieties():
    varieties = BerryVariety.query.all()
    return render_template('varieties.html', varieties=varieties)

@app.route('/add_variety', methods=['GET', 'POST'])
@login_required
def add_variety():
    form = BerryVarietyForm()
    if form.validate_on_submit():
        v = BerryVariety(name=form.name.data, piece_rate_modifier=float(form.piece_rate_modifier.data))
        db.session.add(v)
        db.session.commit()
        flash('Dodano odmianę.', 'success')
        return redirect(url_for('varieties'))
    return render_template('add_variety.html', form=form)

@app.route('/edit_variety/<int:variety_id>', methods=['GET', 'POST'])
@login_required
def edit_variety(variety_id):
    variety = BerryVariety.query.get_or_404(variety_id)
    form = BerryVarietyForm(obj=variety)
    if form.validate_on_submit():
        variety.name = form.name.data
        variety.piece_rate_modifier = float(form.piece_rate_modifier.data)
        db.session.commit()
        flash('Zaktualizowano odmianę.', 'success')
        return redirect(url_for('varieties'))
    return render_template('edit_variety.html', form=form, variety=variety)

@app.route('/delete_variety/<int:variety_id>', methods=['POST'])
@login_required
def delete_variety(variety_id):
    variety = BerryVariety.query.get_or_404(variety_id)
    db.session.delete(variety)
    db.session.commit()
    flash('Usunięto odmianę.', 'success')
    return redirect(url_for('varieties'))

# --- TYPY PRACY ---
@app.route('/work_types')
@login_required
def work_types():
    work_types = WorkType.query.all()
    return render_template('work_types.html', work_types=work_types)

@app.route('/add_work_type', methods=['GET', 'POST'])
@login_required
def add_work_type():
    form = WorkTypeForm()
    if form.validate_on_submit():
        wt = WorkType(
            name=form.name.data,
            is_piece_rate=form.is_piece_rate.data,
            unit=form.unit.data
        )
        db.session.add(wt)
        db.session.commit()
        flash('Dodano typ pracy.', 'success')
        return redirect(url_for('work_types'))
    return render_template('add_work_type.html', form=form)

@app.route('/edit_work_type/<int:work_type_id>', methods=['GET', 'POST'])
@login_required
def edit_work_type(work_type_id):
    wt = WorkType.query.get_or_404(work_type_id)
    form = WorkTypeForm(obj=wt)
    if form.validate_on_submit():
        wt.name = form.name.data
        wt.is_piece_rate = form.is_piece_rate.data
        wt.unit = form.unit.data
        db.session.commit()
        flash('Zaktualizowano typ pracy.', 'success')
        return redirect(url_for('work_types'))
    return render_template('edit_work_type.html', form=form, work_type=wt)

@app.route('/delete_work_type/<int:work_type_id>', methods=['POST'])
@login_required
def delete_work_type(work_type_id):
    wt = WorkType.query.get_or_404(work_type_id)
    db.session.delete(wt)
    db.session.commit()
    flash('Usunięto typ pracy.', 'success')
    return redirect(url_for('work_types'))

# --- ZBIORY DZIENNE ---
@app.route('/add_daily_harvest', methods=['GET', 'POST'])
@login_required
def add_daily_harvest():
    employees = Employee.query.all()
    varieties = BerryVariety.query.all()
    fields = Field.query.all()
    today_date = date.today().isoformat()

    if request.method == 'POST':
        date_selected = request.form['date']
        employee_id = int(request.form['employee_id'])
        quantity_kg = float(request.form['quantity_kg'])
        variety_id = int(request.form['variety_id'])
        field_id = int(request.form['field_id'])
        comment = request.form.get('comment', '')
        piece_rate = request.form.get('piece_rate')

        if not piece_rate:
            employee = Employee.query.get(employee_id)
            piece_rate = employee.piece_rate
        else:
            piece_rate = float(piece_rate)

        date_obj = datetime.strptime(date_selected, "%Y-%m-%d").date()

        harvest = DailyHarvest(
            date=date_obj,
            employee_id=employee_id,
            quantity_kg=quantity_kg,
            variety_id=variety_id,
            field_id=field_id,
            comment=comment
        )
        db.session.add(harvest)
        db.session.commit()

        worktype = WorkType.query.filter_by(is_piece_rate=True).first()
        if not worktype:
            flash('Brak zdefiniowanego typu pracy "akordowej". Dodaj taki typ w "Typy pracy"!', 'danger')
            return redirect(url_for('add_daily_harvest'))

        entry = Entry(
            date=date_obj,
            employee_id=employee_id,
            work_type_id=worktype.id,
            hours=0,
            quantity=quantity_kg,
            variety_id=variety_id,
            field_id=field_id,
            comment='(Automatyczny wpis z dziennego zbioru) ' + (comment or ''),
            piece_rate=piece_rate
        )
        db.session.add(entry)
        db.session.commit()

        flash('Dodano zbiór i automatyczny wpis pracy (akord).', 'success')
        return redirect(url_for('harvests'))

    return render_template(
        'add_daily_harvest.html',
        employees=employees,
        varieties=varieties,
        fields=fields,
        today_date=today_date
    )

@app.route('/harvests')
@login_required
def harvests():
    harvests = DailyHarvest.query.order_by(DailyHarvest.date.desc()).all()
    return render_template('daily_harvest_list.html', harvests=harvests)

@app.route('/clear_day', methods=['POST'])
@login_required
def clear_day():
    date_str = request.form.get('clear_date')
    if not date_str:
        flash('Nie podano daty!', 'danger')
        return redirect(url_for('harvests'))
    try:
        clear_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Błędny format daty!', 'danger')
        return redirect(url_for('harvests'))
    entry_count = Entry.query.filter(Entry.date == clear_date).delete(synchronize_session=False)
    harvest_count = DailyHarvest.query.filter(DailyHarvest.date == clear_date).delete(synchronize_session=False)
    db.session.commit()
    flash(f"Usunięto {harvest_count} zbiorów i {entry_count} wpisów pracy dla dnia {clear_date}.", "success")
    return redirect(url_for('harvests'))

@app.route('/edit_daily_harvest/<int:harvest_id>', methods=['GET', 'POST'])
@login_required
def edit_daily_harvest(harvest_id):
    harvest = DailyHarvest.query.get_or_404(harvest_id)
    form = DailyHarvestForm(obj=harvest)
    form.employee_id.choices = [(e.id, e.name) for e in Employee.query.all()]
    form.variety_id.choices = [(v.id, v.name) for v in BerryVariety.query.all()]
    form.field_id.choices = [(f.id, f.name) for f in Field.query.all()]
    if form.validate_on_submit():
        harvest.date = form.date.data
        harvest.employee_id = form.employee_id.data
        harvest.quantity_kg = float(form.quantity_kg.data)
        harvest.variety_id = form.variety_id.data
        harvest.field_id = form.field_id.data
        harvest.comment = form.comment.data
        db.session.commit()
        flash('Zaktualizowano zbiór.', 'success')
        return redirect(url_for('harvests'))
    return render_template('edit_daily_harvest.html', form=form)

@app.route('/delete_daily_harvest/<int:harvest_id>', methods=['POST'])
@login_required
def delete_daily_harvest(harvest_id):
    harvest = DailyHarvest.query.get_or_404(harvest_id)
    related_entries = Entry.query.filter_by(
        date=harvest.date,
        employee_id=harvest.employee_id,
        variety_id=harvest.variety_id,
        field_id=harvest.field_id,
        quantity=harvest.quantity_kg
    ).all()
    for entry in related_entries:
        db.session.delete(entry)
    db.session.delete(harvest)
    db.session.commit()
    flash('Usunięto zbiór i powiązany wpis akordowy.', 'success')
    return redirect(url_for('harvests'))

# --- WPISY PRACY (ENTRY) ---
@app.route('/entries')
@login_required
def entries():
    employees = Employee.query.all()
    work_types = WorkType.query.all()
    selected_employee = request.args.get('employee_filter', type=int)
    selected_work_type = request.args.get('work_type_filter', type=int)
    start_date_filter = request.args.get('start_date_filter')
    end_date_filter = request.args.get('end_date_filter')

    query = Entry.query
    if selected_employee:
        query = query.filter_by(employee_id=selected_employee)
    if selected_work_type:
        query = query.filter_by(work_type_id=selected_work_type)
    if start_date_filter:
        query = query.filter(Entry.date >= start_date_filter)
    if end_date_filter:
        query = query.filter(Entry.date <= end_date_filter)

    entries = query.order_by(Entry.date.desc()).all()
    for entry in entries:
        entry.employee_obj = entry.employee
        entry.work_type_obj = entry.work_type
        entry.variety_obj = entry.variety if entry.variety else None
        entry.field_obj = entry.field if entry.field else None

    return render_template(
        'entries.html',
        entries=entries,
        employees=employees,
        work_types=work_types,
        selected_employee=selected_employee,
        selected_work_type=selected_work_type,
        start_date_filter=start_date_filter,
        end_date_filter=end_date_filter
    )

@app.route('/add_entry', methods=['GET', 'POST'])
@login_required
def add_entry():
    employees = Employee.query.all()
    work_types = WorkType.query.all()
    varieties = BerryVariety.query.all()
    fields = Field.query.all()
    today_date = date.today().isoformat()
    if request.method == 'POST':
        date_selected = request.form['date']
        employee_id = int(request.form['employee_id'])
        work_type_id = int(request.form['work_type_id'])
        hours = float(request.form.get('hours', 0) or 0)
        quantity = float(request.form.get('quantity', 0) or 0)
        variety_id = request.form.get('variety_id')
        field_id = request.form.get('field_id')
        comment = request.form.get('comment', '')
        piece_rate = request.form.get('piece_rate')

        if not piece_rate:
            employee = Employee.query.get(employee_id)
            piece_rate = employee.piece_rate
        else:
            piece_rate = float(piece_rate)

        entry_date = datetime.strptime(date_selected, '%Y-%m-%d').date()

        entry = Entry(
            date=entry_date,
            employee_id=employee_id,
            work_type_id=work_type_id,
            hours=hours,
            quantity=quantity,
            variety_id=int(variety_id) if variety_id else None,
            field_id=int(field_id) if field_id else None,
            comment=comment,
            piece_rate=float(piece_rate)
        )
        db.session.add(entry)
        db.session.commit()
        flash('Dodano wpis pracy.', 'success')
        return redirect(url_for('entries'))
    return render_template('add_entry.html', employees=employees, work_types=work_types, varieties=varieties, fields=fields, today_date=today_date)

@app.route('/edit_entry/<int:entry_id>', methods=['GET', 'POST'])
@login_required
def edit_entry(entry_id):
    entry = Entry.query.get_or_404(entry_id)
    employees = Employee.query.all()
    work_types = WorkType.query.all()
    varieties = BerryVariety.query.all()
    fields = Field.query.all()
    if request.method == 'POST':
        date_selected = request.form['date']
        employee_id = int(request.form['employee_id'])
        work_type_id = int(request.form['work_type_id'])
        hours = float(request.form.get('hours', 0) or 0)
        quantity = float(request.form.get('quantity', 0) or 0)
        variety_id = request.form.get('variety_id')
        field_id = request.form.get('field_id')
        comment = request.form.get('comment', '')
        piece_rate = request.form.get('piece_rate')

        entry.date = datetime.strptime(date_selected, '%Y-%m-%d').date()
        entry.employee_id = employee_id
        entry.work_type_id = work_type_id
        entry.hours = hours
        entry.quantity = quantity
        entry.variety_id = int(variety_id) if variety_id else None
        entry.field_id = int(field_id) if field_id else None
        entry.comment = comment

        if not piece_rate:
            employee = Employee.query.get(employee_id)
            piece_rate = employee.piece_rate
        else:
            piece_rate = float(piece_rate)
        entry.piece_rate = float(piece_rate)

        db.session.commit()
        flash('Zaktualizowano wpis pracy.', 'success')
        return redirect(url_for('entries'))
    return render_template('edit_entry.html', entry=entry, employees=employees, work_types=work_types, varieties=varieties, fields=fields)

@app.route('/delete_entry/<int:entry_id>', methods=['POST'])
@login_required
def delete_entry(entry_id):
    entry = Entry.query.get_or_404(entry_id)
    db.session.delete(entry)
    db.session.commit()
    flash('Usunięto wpis pracy.', 'success')
    return redirect(url_for('entries'))

# --- RAPORTY ---
@app.route('/employee_reports', methods=['GET'])
@login_required
def employee_reports():
    employees = Employee.query.order_by(Employee.name).all()
    selected_employee_ids = request.args.getlist('employee_id')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not selected_employee_ids or '' in selected_employee_ids:
        selected_emps = employees
        selected_employee_ids = []
    else:
        selected_emps = Employee.query.filter(Employee.id.in_(selected_employee_ids)).all()

    summaries = []
    if selected_emps and start_date and end_date:
        for employee in selected_emps:
            hours = db.session.query(db.func.sum(Entry.hours)).filter(
                Entry.employee_id == employee.id,
                Entry.date >= start_date,
                Entry.date <= end_date,
                Entry.hours > 0
            ).scalar() or 0
            piece_entries = Entry.query.filter(
                Entry.employee_id == employee.id,
                Entry.date >= start_date,
                Entry.date <= end_date,
                Entry.quantity > 0
            ).all()
            quantity = sum(e.quantity or 0 for e in piece_entries)
            piece_pay = 0
            for e in piece_entries:
                modifier = 1.0
                if e.variety and hasattr(e.variety, "piece_rate_modifier"):
                    modifier = e.variety.piece_rate_modifier
                piece_pay += (e.quantity or 0) * (e.piece_rate if e.piece_rate is not None else employee.piece_rate) * modifier
            summaries.append({
                'employee': employee,
                'hours': hours,
                'hourly_pay': round(hours * employee.hourly_rate, 2),
                'quantity': quantity,
                'piece_pay': round(piece_pay, 2),
                'start_date': start_date,
                'end_date': end_date
            })
    else:
        summaries = None

    return render_template(
        'employee_reports.html',
        employees=employees,
        summaries=summaries,
        selected_employee_ids=selected_employee_ids,
        start_date=start_date,
        end_date=end_date
    )

@app.route('/employee_reports_export')
@login_required
def employee_reports_export():
    selected_employee_ids = request.args.getlist('employee_id')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    employees = Employee.query.order_by(Employee.name).all()
    if not selected_employee_ids or '' in selected_employee_ids:
        selected_emps = employees
    else:
        selected_emps = Employee.query.filter(Employee.id.in_(selected_employee_ids)).all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Raport pracowników"

    headers = [
        "Pracownik",
        "Okres od",
        "Okres do",
        "Suma godzin",
        "Wynagrodzenie godzinowe",
        "Suma zbiorów (akord)",
        "Wynagrodzenie akordowe"
    ]
    sheet.append(headers)

    for employee in selected_emps:
        hours = db.session.query(db.func.sum(Entry.hours)).filter(
            Entry.employee_id == employee.id,
            Entry.date >= start_date,
            Entry.date <= end_date,
            Entry.hours > 0
        ).scalar() or 0

        piece_entries = Entry.query.filter(
            Entry.employee_id == employee.id,
            Entry.date >= start_date,
            Entry.date <= end_date,
            Entry.quantity > 0
        ).all()
        quantity = sum(e.quantity or 0 for e in piece_entries)
        piece_pay = 0
        for e in piece_entries:
            modifier = 1.0
            if e.variety and hasattr(e.variety, "piece_rate_modifier"):
                modifier = e.variety.piece_rate_modifier
            piece_pay += (e.quantity or 0) * (e.piece_rate if e.piece_rate is not None else employee.piece_rate) * modifier

        hourly_pay = round(hours * employee.hourly_rate, 2)

        row = [
            employee.name,
            start_date,
            end_date,
            "%.2f" % hours,
            "%.2f PLN" % hourly_pay,
            "%.2f" % quantity,
            "%.2f PLN" % piece_pay
        ]
        sheet.append(row)

    last_row = sheet.max_row
    sum_hours = sum(float(sheet.cell(row=i, column=4).value) for i in range(2, last_row + 1))
    sum_hourly_pay = sum(float(str(sheet.cell(row=i, column=5).value).replace(" PLN", "")) for i in range(2, last_row + 1))
    sum_quantity = sum(float(sheet.cell(row=i, column=6).value) for i in range(2, last_row + 1))
    sum_piece_pay = sum(float(str(sheet.cell(row=i, column=7).value).replace(" PLN", "")) for i in range(2, last_row + 1))

    sheet.append([
        "SUMA:", "", "",
        "%.2f" % sum_hours,
        "%.2f PLN" % sum_hourly_pay,
        "%.2f" % sum_quantity,
        "%.2f PLN" % sum_piece_pay
    ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="raport_pracownikow.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/daily_report', methods=['GET'])
@login_required
def daily_report():
    employees = Employee.query.order_by(Employee.name).all()
    selected_date = request.args.get('date')
    data = []

    if selected_date:
        selected_date_dt = datetime.strptime(selected_date, "%Y-%m-%d").date()
    else:
        selected_date_dt = date.today()

    for emp in employees:
        hours = db.session.query(db.func.sum(Entry.hours)).filter(
            Entry.employee_id == emp.id,
            Entry.date == selected_date_dt,
            Entry.hours > 0
        ).scalar() or 0

        hourly_pay = round(hours * emp.hourly_rate, 2)

        entries_piece = db.session.query(Entry).filter(
            Entry.employee_id == emp.id,
            Entry.date == selected_date_dt,
            Entry.quantity > 0
        ).all()
        quantity = sum(e.quantity or 0 for e in entries_piece)
        piece_pay = 0
        for e in entries_piece:
            modifier = 1.0
            if e.variety and hasattr(e.variety, "piece_rate_modifier"):
                modifier = e.variety.piece_rate_modifier
            piece_pay += (e.quantity or 0) * (e.piece_rate if e.piece_rate is not None else emp.piece_rate) * modifier
        piece_pay = round(piece_pay, 2)

        data.append({
            'employee': emp,
            'hours': hours,
            'hourly_pay': hourly_pay,
            'quantity': quantity,
            'piece_pay': piece_pay
        })

    return render_template(
        'daily_report.html',
        data=data,
        selected_date=selected_date_dt.strftime('%Y-%m-%d')
    )

@app.route('/generate_report', methods=['GET', 'POST'])
@login_required
def generate_report():
    employees = Employee.query.all()
    if request.method == 'POST':
        employee_id = request.form.get('employee_id')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')

        query = Entry.query
        if employee_id:
            query = query.filter_by(employee_id=employee_id)
        if start_date:
            query = query.filter(Entry.date >= start_date)
        if end_date:
            query = query.filter(Entry.date <= end_date)
        entries = query.order_by(Entry.date).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Wpisy pracy"
        ws.append(['Data', 'Pracownik', 'Typ pracy', 'Godziny', 'Ilość', 'Odmiana', 'Pole', 'Komentarz'])
        for entry in entries:
            ws.append([
                entry.date.strftime('%Y-%m-%d'),
                entry.employee.name if entry.employee else '',
                entry.work_type.name if entry.work_type else '',
                entry.hours if entry.hours else '',
                entry.quantity if entry.quantity else '',
                entry.variety.name if entry.variety else '',
                entry.field.name if entry.field else '',
                entry.comment or ''
            ])
        f = BytesIO()
        wb.save(f)
        f.seek(0)
        return send_file(
            f,
            as_attachment=True,
            download_name="wpisy_pracy.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    return render_template('generate_report.html', employees=employees)

@app.route('/harvest_overview', methods=['GET'])
def harvest_overview():
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.name).all()
    selected_id = request.args.get('employee_id', type=int)
    selected_employee = None
    summary = []

    if selected_id:
        selected_employee = Employee.query.get(selected_id)
        all_dates = set(
            [h.date for h in DailyHarvest.query.filter_by(employee_id=selected_id).all()] +
            [e.date for e in Entry.query.filter_by(employee_id=selected_id).all()]
        )
        for d in sorted(all_dates, reverse=True):
            harvests = DailyHarvest.query.filter_by(employee_id=selected_id, date=d).all()
            total_kg = sum(h.quantity_kg for h in harvests)
            odmiany = ', '.join(sorted(set(h.variety.name for h in harvests if h.variety)))
            pola = ', '.join(sorted(set(h.field.name for h in harvests if h.field)))
            total_piece_pay = 0
            for h in harvests:
                modifier = 1.0
                if h.variety and hasattr(h.variety, "piece_rate_modifier"):
                    modifier = h.variety.piece_rate_modifier
                rate = selected_employee.piece_rate or 0
                total_piece_pay += (h.quantity_kg or 0) * rate * modifier

            entries = Entry.query.filter_by(employee_id=selected_id, date=d).all()
            total_hours = sum(e.hours or 0 for e in entries if (e.hours or 0) > 0)
            hourly_pay = (selected_employee.hourly_rate or 0) * total_hours

            summary.append({
                'date': d.strftime('%Y-%m-%d'),
                'odmiany': odmiany,
                'pola': pola,
                'total_kg': round(total_kg, 2),
                'total_piece_pay': round(total_piece_pay, 2),
                'total_hours': round(total_hours, 2),
                'hourly_pay': round(hourly_pay, 2),
                'day_sum': round(total_piece_pay + hourly_pay, 2)
            })

    total_kg_sum = sum(row['total_kg'] for row in summary)
    total_piece_sum = sum(row['total_piece_pay'] for row in summary)
    total_hours_sum = sum(row['total_hours'] for row in summary)
    total_hourly_sum = sum(row['hourly_pay'] for row in summary)
    total_all_sum = sum(row['day_sum'] for row in summary)

    return render_template(
        'harvest_overview.html',
        employees=employees,
        selected_employee=selected_employee,
        summary=summary,
        total_kg_sum=round(total_kg_sum, 2),
        total_piece_sum=round(total_piece_sum, 2),
        total_hours_sum=round(total_hours_sum, 2),
        total_hourly_sum=round(total_hourly_sum, 2),
        total_all_sum=round(total_all_sum, 2),
        selected_id=selected_id
    )

@app.route('/fast_entry', methods=['GET', 'POST'])
@login_required
def fast_entry():
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.name).all()
    fields = Field.query.order_by(Field.name).all()
    varieties = BerryVariety.query.order_by(BerryVariety.name).all()
    today = date.today().isoformat()
    success = False

    if request.method == 'POST':
        date_selected = request.form['data']
        employee_id = int(request.form['employee_id'])
        field_id = int(request.form['field_id'])
        variety_id = int(request.form['variety_id'])
        waga = float(request.form['waga'])

        # domyślny piece_rate
        employee = Employee.query.get(employee_id)
        piece_rate = employee.piece_rate

        # Dodaj rekord zbioru
        harvest = DailyHarvest(
            date=datetime.strptime(date_selected, '%Y-%m-%d').date(),
            employee_id=employee_id,
            quantity_kg=waga,
            variety_id=variety_id,
            field_id=field_id,
            comment="Szybki wpis"
        )
        db.session.add(harvest)
        db.session.commit()

        # Dodaj automatycznie wpis pracy (jeśli typ akordowy istnieje)
        worktype = WorkType.query.filter_by(is_piece_rate=True).first()
        if worktype:
            entry = Entry(
                date=datetime.strptime(date_selected, '%Y-%m-%d').date(),
                employee_id=employee_id,
                work_type_id=worktype.id,
                hours=0,
                quantity=waga,
                variety_id=variety_id,
                field_id=field_id,
                comment='(Automatyczny szybki wpis zbioru)',
                piece_rate=piece_rate
            )
            db.session.add(entry)
            db.session.commit()

        success = True

    # Pokaż 5 ostatnich wpisów
    last_entries = (
        DailyHarvest.query
        .order_by(DailyHarvest.id.desc())
        .limit(5)
        .all()
    )

    return render_template(
        'fast_entry.html',
        employees=employees,
        fields=fields,
        varieties=varieties,
        today=today,
        success=success,
        last_entries=last_entries
    )

from flask_migrate import upgrade
with app.app_context():
    upgrade()